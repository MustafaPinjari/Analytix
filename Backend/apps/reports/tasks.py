from celery import shared_task
import os
import json
import logging
from django.core.mail import EmailMessage
from django.conf import settings
from apps.reports.models import Report, ReportHistory
from openpyxl import Workbook
import uuid

logger = logging.getLogger(__name__)

@shared_task
def compile_report_task(report_id: str):
    logger.info(f"Starting report compile task for report_id: {report_id}")
    try:
        report = Report.objects.get(id=report_id)
    except Report.DoesNotExist:
        logger.error(f"Report schedule {report_id} does not exist.")
        return

    history = ReportHistory.objects.create(
        report=report,
        status="RUNNING"
    )

    try:
        # Resolve target files directory
        reports_dir = os.path.join(settings.MEDIA_ROOT, "reports")
        os.makedirs(reports_dir, exist_ok=True)
        
        filename = f"{report.name.strip().replace(' ', '_')}_{uuid.uuid4().hex[:8]}"
        
        if report.format == "EXCEL":
            filename = f"{filename}.xlsx"
            file_path = os.path.join(reports_dir, filename)
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Report Details"
            ws["A1"] = "InsightFlow BI Scheduled Report"
            ws["A2"] = f"Report Name: {report.name}"
            ws["A3"] = f"Format: {report.format}"
            ws["A4"] = f"Dashboard Reference ID: {report.dashboard_id}"
            
            wb.save(file_path)
        else:
            filename = f"{filename}.pdf"
            file_path = os.path.join(reports_dir, filename)
            with open(file_path, "w") as f:
                f.write(f"InsightFlow BI PDF Report\nName: {report.name}\nDashboard ID: {report.dashboard_id}\n")

        recipients_list = json.loads(report.recipients)

        email = EmailMessage(
            subject=f"InsightFlow BI Scheduled Report: {report.name}",
            body=f"Hello,\n\nPlease find attached the scheduled report: {report.name}.\n\nBest regards,\nInsightFlow BI Team",
            from_email="reports@insightflow.bi",
            to=recipients_list
        )
        email.attach_file(file_path)
        email.send(fail_silently=False)

        history.status = "COMPLETED"
        history.file_path = file_path
        history.save()
        logger.info(f"Report compile task succeeded for report_id: {report_id}")
    except Exception as e:
        logger.error(f"Failed to generate report {report_id}: {str(e)}")
        history.status = "FAILED"
        history.error_message = str(e)
        history.save()
